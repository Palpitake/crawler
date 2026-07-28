"""Deterministic Code artifact inspection helpers.

The Code Agent itself uses pi-coding-agent native read/write/edit/bash tools.
This module only validates source and output artifacts after the native session.
"""

from __future__ import annotations

import ast
import csv
import json
import os
import re
import subprocess
import sys
import time
import uuid
from pathlib import Path
from string import Template
from typing import Any, Dict, List


_WORKSPACE = Path("./crawler_workspace").resolve()


SENSITIVE_SUFFIXES = {
    ".env",
    ".pem",
    ".key",
    ".crt",
    ".p12",
    ".pfx",
}


SENSITIVE_NAMES = {
    ".env",
    "id_rsa",
    "id_dsa",
    "credentials.json",
    "token.json",
}


def set_workspace(workspace: str) -> None:
    """
    设置 Agent 工作区。

    code_agent.py 会在启动时调用该函数。
    """
    global _WORKSPACE
    _WORKSPACE = Path(workspace).expanduser().resolve()
    _WORKSPACE.mkdir(parents=True, exist_ok=True)


def get_workspace() -> Path:
    """
    获取当前工作区路径。

    code_agent.py 会用它打印工作区和日志文件路径。
    """
    _WORKSPACE.mkdir(parents=True, exist_ok=True)
    return _WORKSPACE


def _json(data: Dict[str, Any]) -> str:
    return json.dumps(data, ensure_ascii=False, indent=2)


def _redact(text: str) -> str:
    """
    简单脱敏，避免工具结果中泄露 token / key / cookie。
    """
    if not text:
        return text

    patterns = [
        (r"(?i)(api[_-]?key\s*[:=]\s*)[^\s]+", r"\1***REDACTED***"),
        (r"(?i)(token\s*[:=]\s*)[^\s]+", r"\1***REDACTED***"),
        (r"(?i)(secret\s*[:=]\s*)[^\s]+", r"\1***REDACTED***"),
        (r"(?i)(password\s*[:=]\s*)[^\s]+", r"\1***REDACTED***"),
        (r"(?i)(cookie\s*[:=]\s*)[^\n]+", r"\1***REDACTED***"),
        (r"sk-[A-Za-z0-9_\-]{10,}", "sk-***REDACTED***"),
    ]

    redacted = text
    for pattern, replacement in patterns:
        redacted = re.sub(pattern, replacement, redacted)

    return redacted


def _safe_path(filename: str) -> Path:
    """
    将文件名转换为工作区内的安全路径。

    禁止：
    - 绝对路径逃逸；
    - ../ 路径逃逸；
    - 读取或写入敏感文件。
    """
    if not filename or not filename.strip():
        raise ValueError("filename 不能为空")

    raw = Path(filename)

    if raw.is_absolute():
        raise ValueError("不允许使用绝对路径")

    path = (get_workspace() / raw).resolve()

    workspace = get_workspace().resolve()

    if workspace not in path.parents and path != workspace:
        raise ValueError("路径逃逸：文件必须位于工作区内")

    if path.name in SENSITIVE_NAMES:
        raise ValueError(f"不允许访问敏感文件：{path.name}")

    if path.suffix.lower() in SENSITIVE_SUFFIXES:
        raise ValueError(f"不允许访问敏感后缀文件：{path.suffix}")

    return path


def _file_info(path: Path) -> Dict[str, Any]:
    stat = path.stat()
    return {
        "name": path.name,
        "relative_path": str(path.relative_to(get_workspace())),
        "path": str(path),
        "size_bytes": stat.st_size,
        "modified_time": time.strftime(
            "%Y-%m-%d %H:%M:%S",
            time.localtime(stat.st_mtime),
        ),
    }


def read_text_file(
    filename: str,
    max_chars: int = 20000,
    start_line: int = 1,
    end_line: int | None = None,
) -> str:
    """
    按行读取工作区内的文本文件内容。
    start_line/end_line 均为从 1 开始的闭区间；可分段读取长代码。
    """
    try:
        path = _safe_path(filename)

        if not path.exists():
            return _json({
                "success": False,
                "error": "文件不存在",
                "filename": filename,
            })

        if not path.is_file():
            return _json({
                "success": False,
                "error": "目标不是文件",
                "filename": filename,
            })

        text = path.read_text(encoding="utf-8", errors="replace")
        lines = text.splitlines(keepends=True)
        total_lines = len(lines)
        try:
            start = max(1, int(start_line or 1))
        except Exception:
            start = 1
        try:
            requested_end = total_lines if end_line is None else min(total_lines, int(end_line))
        except Exception:
            requested_end = total_lines
        requested_end = max(start - 1, requested_end)
        max_chars = max(1000, min(int(max_chars or 20000), 100000))

        selected: List[str] = []
        char_count = 0
        actual_end = start - 1
        for line_number in range(start, requested_end + 1):
            if line_number > total_lines:
                break
            line = lines[line_number - 1]
            if selected and char_count + len(line) > max_chars:
                break
            if not selected and len(line) > max_chars:
                line = line[:max_chars]
            selected.append(line)
            char_count += len(line)
            actual_end = line_number

        chunk = "".join(selected)
        truncated = actual_end < requested_end or requested_end < total_lines

        return _json({
            "success": True,
            "filename": filename,
            "path": str(path),
            "truncated": truncated,
            "start_line": start,
            "end_line": actual_end,
            "total_lines": total_lines,
            "has_more": actual_end < total_lines,
            "next_start_line": actual_end + 1 if actual_end < total_lines else None,
            "content": _redact(chunk),
        })

    except Exception as e:
        return _json({
            "success": False,
            "error": repr(e),
        })


def _call_name(node: ast.AST) -> str:
    """
    获取函数调用名：
    - os.system -> "os.system"
    - subprocess.run -> "subprocess.run"
    - eval -> "eval"
    """
    if isinstance(node, ast.Name):
        return node.id

    if isinstance(node, ast.Attribute):
        base = _call_name(node.value)
        return f"{base}.{node.attr}" if base else node.attr

    return ""


def _inspect_dangerous_calls(tree: ast.AST) -> List[Dict[str, Any]]:
    dangerous_exact = {
        "eval",
        "exec",
        "compile",
        "__import__",
        "os.system",
        "os.popen",
        "os.fork",
        "os.spawnl",
        "os.spawnle",
        "os.spawnlp",
        "os.spawnlpe",
        "os.spawnv",
        "os.spawnve",
        "os.spawnvp",
        "os.spawnvpe",
        "subprocess.run",
        "subprocess.Popen",
        "subprocess.call",
        "subprocess.check_call",
        "subprocess.check_output",
        "shutil.rmtree",
        "os.getenv",
        "os.environ.get",
        "pathlib.Path.home",
        "Path.home",
    }

    dangerous_suffix = {
        ".unlink",
        ".rmdir",
        ".remove",
    }

    findings: List[Dict[str, Any]] = []

    for node in ast.walk(tree):
        if not isinstance(node, ast.Call):
            continue

        name = _call_name(node.func)

        if name in dangerous_exact or any(name.endswith(s) for s in dangerous_suffix):
            findings.append({
                "line": getattr(node, "lineno", None),
                "call": name,
                "level": "danger",
            })

    return findings


def _inspect_dangerous_imports(tree: ast.AST) -> List[Dict[str, Any]]:
    """生成的爬虫不需要直接导入这些高风险系统模块。"""
    denied = {
        "subprocess", "ctypes", "multiprocessing", "pickle", "marshal",
        "socket", "ftplib", "smtplib", "paramiko",
    }
    findings: List[Dict[str, Any]] = []
    for node in ast.walk(tree):
        names: List[str] = []
        if isinstance(node, ast.Import):
            names = [alias.name.split(".", 1)[0] for alias in node.names]
        elif isinstance(node, ast.ImportFrom) and node.module:
            names = [node.module.split(".", 1)[0]]
        for name in names:
            if name in denied:
                findings.append({
                    "line": getattr(node, "lineno", None),
                    "import": name,
                    "level": "danger",
                })
    return findings


def _has_main_function(tree: ast.AST) -> bool:
    for node in ast.walk(tree):
        if isinstance(node, ast.FunctionDef) and node.name == "main":
            return True
    return False


def _has_main_guard(text: str) -> bool:
    return '__name__ == "__main__"' in text or "__name__ == '__main__'" in text


def _has_save_logic(text: str) -> bool:
    save_markers = [
        ".to_csv(",
        ".to_json(",
        ".to_excel(",
        "json.dump(",
        "csv.writer(",
        "DictWriter(",
        "open(",
    ]

    data_suffix_markers = [
        ".csv",
        ".json",
        ".xlsx",
    ]

    return any(m in text for m in save_markers) and any(s in text for s in data_suffix_markers)


def inspect_python_file(filename: str) -> str:
    """
    检查工作区内 Python 文件的语法、安全性和爬虫代码基本要求。
    """
    try:
        path = _safe_path(filename)

        if path.suffix != ".py":
            return _json({
                "success": False,
                "error": "只能检查 .py 文件",
                "filename": filename,
            })

        if not path.exists():
            return _json({
                "success": False,
                "error": "文件不存在",
                "filename": filename,
            })

        text = path.read_text(encoding="utf-8", errors="replace")

        try:
            tree = ast.parse(text, filename=str(path))
            syntax_ok = True
            syntax_error = None
        except SyntaxError as e:
            return _json({
                "success": False,
                "syntax_ok": False,
                "error": f"语法错误：{e}",
                "line": e.lineno,
                "offset": e.offset,
            })

        dangerous_calls = _inspect_dangerous_calls(tree)
        dangerous_imports = _inspect_dangerous_imports(tree)

        warnings: List[str] = []

        if not _has_main_function(tree):
            warnings.append("未检测到 main() 函数")

        if not _has_main_guard(text):
            warnings.append("未检测到 if __name__ == '__main__' 入口")

        if not _has_save_logic(text):
            warnings.append("未检测到明确的数据保存逻辑，如 CSV / JSON / Excel 保存")

        if "timeout=" not in text:
            warnings.append("未检测到请求 timeout 参数")

        if "User-Agent" not in text and "user-agent" not in text.lower():
            warnings.append("未检测到 User-Agent 设置")

        if "logging" not in text and "logger" not in text:
            warnings.append("未检测到 logging 日志逻辑")

        if "try:" not in text or "except " not in text:
            warnings.append("异常处理不明显，建议增加 try/except")

        if "sleep(" not in text and "time.sleep" not in text and "retry" not in text.lower():
            warnings.append("未检测到 sleep 或 retry，建议礼貌访问并避免请求过快")

        passed = syntax_ok and len(dangerous_calls) == 0 and len(dangerous_imports) == 0

        return _json({
            "success": passed,
            "syntax_ok": syntax_ok,
            "syntax_error": syntax_error,
            "dangerous_calls": dangerous_calls,
            "dangerous_imports": dangerous_imports,
            "warnings": warnings,
            "message": "检查通过" if passed else "检查未通过：存在危险调用或严重问题",
        })

    except Exception as e:
        return _json({
            "success": False,
            "error": repr(e),
        })


def create_debug_python_file(
    source_filename: str,
    error_context: Any = "",
    focus: str = "",
    context_lines: int = 3,
    max_snippets: int = 12,
) -> Dict[str, Any]:
    """Create a unique stdlib-only Python probe for one source file."""
    source_path = _safe_path(source_filename)
    if source_path.suffix.lower() != ".py":
        raise ValueError("只能为 .py 源文件创建 Python Debug 脚本")
    source_relative = str(source_path.relative_to(get_workspace()))
    safe_stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", source_path.stem).strip("._") or "code"
    stamp = time.strftime("%Y%m%d_%H%M%S", time.localtime())
    debug_id = f"{stamp}_{time.time_ns() % 1_000_000_000:09d}_{uuid.uuid4().hex[:8]}"
    debug_relative = Path("debug") / f"{safe_stem}.debug.{debug_id}.py"
    debug_path = _safe_path(str(debug_relative))
    debug_path.parent.mkdir(parents=True, exist_ok=True)

    focus_terms = [term.strip().lower() for term in re.split(r"[,\s]+", str(focus or "")) if term.strip()]
    if not focus_terms:
        focus_terms = [
            "request_timeout_seconds", "max_stalled_cursors", "seen_cursors",
            "next_cursor", "has_more", "crawl_progress_json", "crawl_meta_json",
            "expect_response", "wait_for_response", "response", "except",
        ]
    try:
        context = max(0, min(int(context_lines or 3), 10))
    except Exception:
        context = 3
    try:
        snippet_limit = max(1, min(int(max_snippets or 12), 30))
    except Exception:
        snippet_limit = 12

    if isinstance(error_context, str):
        context_value: Any = _redact(error_context)[:12000]
    else:
        context_value = _redact(json.dumps(error_context, ensure_ascii=False, default=str))[:12000]

    template = r'''"""Generated code-debug probe.

Run without arguments for safe static diagnostics.
Run with ``--run-target`` only when reproducing the target's runtime failure is intended.
"""
from __future__ import annotations

import ast
import faulthandler
import json
import runpy
import sys
import traceback
from pathlib import Path

DEBUG_MARKER = "CODE_DEBUG_JSON="
WORKSPACE = Path(__file__).resolve().parent.parent
TARGET_RELATIVE = $TARGET_RELATIVE
TARGET = (WORKSPACE / TARGET_RELATIVE).resolve()
ERROR_CONTEXT = $ERROR_CONTEXT
FOCUS_TERMS = $FOCUS_TERMS
CONTEXT_LINES = $CONTEXT_LINES
MAX_SNIPPETS = $MAX_SNIPPETS


def static_diagnostics() -> dict:
    result = {
        "target_file": TARGET_RELATIVE,
        "target_path": str(TARGET),
        "exists": TARGET.is_file(),
        "syntax_ok": False,
        "original_error": ERROR_CONTEXT,
    }
    if WORKSPACE not in TARGET.parents or not TARGET.is_file():
        result["error"] = "target file missing or outside workspace"
        return result

    text = TARGET.read_text(encoding="utf-8", errors="replace")
    lines = text.splitlines()
    result.update({
        "source_sha256": hashlib.sha256(text.encode("utf-8")).hexdigest(),
        "size_chars": len(text),
        "total_lines": len(lines),
    })
    try:
        compile(text, str(TARGET), "exec")
        tree = ast.parse(text, filename=str(TARGET))
    except SyntaxError as exc:
        line_no = int(exc.lineno or 1)
        start = max(1, line_no - 4)
        end = min(len(lines), line_no + 4)
        result.update({
            "error_type": "SyntaxError",
            "error": str(exc),
            "line": line_no,
            "offset": exc.offset,
            "snippet": "\n".join(
                f"{number:04d}: {lines[number - 1]}"
                for number in range(start, end + 1)
            ),
        })
        return result

    result["syntax_ok"] = True
    outline = []
    constants = {}
    for node in tree.body:
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef, ast.ClassDef)):
            outline.append({
                "kind": type(node).__name__,
                "name": node.name,
                "start_line": getattr(node, "lineno", None),
                "end_line": getattr(node, "end_lineno", None),
            })
        elif isinstance(node, (ast.Assign, ast.AnnAssign)):
            targets = node.targets if isinstance(node, ast.Assign) else [node.target]
            value = getattr(node, "value", None)
            if isinstance(value, ast.Constant) and isinstance(value.value, (str, int, float, bool, type(None))):
                for target in targets:
                    if isinstance(target, ast.Name) and target.id.isupper():
                        constants[target.id] = value.value

    matched = [
        number for number, line in enumerate(lines, start=1)
        if any(term in line.lower() for term in FOCUS_TERMS)
    ]
    snippets = []
    covered_until = 0
    for line_no in matched:
        if len(snippets) >= MAX_SNIPPETS:
            break
        start = max(1, line_no - CONTEXT_LINES)
        end = min(len(lines), line_no + CONTEXT_LINES)
        if start <= covered_until:
            continue
        snippets.append({
            "start_line": start,
            "end_line": end,
            "text": "\n".join(
                f"{number:04d}: {lines[number - 1]}"
                for number in range(start, end + 1)
            ),
        })
        covered_until = end
    result.update({
        "outline": outline[:80],
        "constants": constants,
        "focus_terms": FOCUS_TERMS,
        "matched_line_count": len(matched),
        "snippets": snippets,
    })
    return result


def main() -> int:
    faulthandler.enable()
    report = {
        "debug_file": str(Path(__file__).relative_to(WORKSPACE)),
        "mode": "runtime" if "--run-target" in sys.argv else "static",
        "static": static_diagnostics(),
    }
    if "--run-target" in sys.argv and report["static"].get("syntax_ok"):
        original_argv = sys.argv[:]
        sys.argv = [str(TARGET)]
        try:
            runpy.run_path(str(TARGET), run_name="__main__")
            report["runtime"] = {"success": True}
        except SystemExit as exc:
            code = exc.code if isinstance(exc.code, int) else 1
            report["runtime"] = {
                "success": code == 0,
                "error_type": "SystemExit",
                "error": str(exc),
                "exit_code": code,
                "traceback": traceback.format_exc()[-12000:],
            }
        except BaseException as exc:
            report["runtime"] = {
                "success": False,
                "error_type": type(exc).__name__,
                "error": str(exc),
                "traceback": traceback.format_exc()[-12000:],
            }
        finally:
            sys.argv = original_argv
    print(DEBUG_MARKER + json.dumps(report, ensure_ascii=False, default=str), flush=True)
    static_ok = bool(report["static"].get("syntax_ok"))
    runtime_ok = bool((report.get("runtime") or {}).get("success", True))
    return 0 if static_ok and runtime_ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
'''
    script = Template(template).substitute(
        TARGET_RELATIVE=repr(source_relative),
        ERROR_CONTEXT=repr(context_value),
        FOCUS_TERMS=repr(focus_terms),
        CONTEXT_LINES=str(context),
        MAX_SNIPPETS=str(snippet_limit),
    )
    debug_path.write_text(script, encoding="utf-8")
    return {
        "success": True,
        "debug_id": debug_id,
        "debug_file": str(debug_relative),
        "target_file": source_relative,
        "file": _file_info(debug_path),
    }


def debug_python_file(
    filename: str,
    focus: str = "",
    error_context: str = "",
    context_lines: int = 3,
    max_snippets: int = 12,
) -> str:
    """创建并运行独立 Python 调试脚本，返回目标代码的静态调试信息。"""
    try:
        created = create_debug_python_file(
            source_filename=filename,
            error_context=error_context,
            focus=focus,
            context_lines=context_lines,
            max_snippets=max_snippets,
        )
        debug_path = _safe_path(str(created["debug_file"]))
        env = dict(os.environ)
        env["PYTHONUNBUFFERED"] = "1"
        env["PYTHONUTF8"] = "1"
        env["PYTHONIOENCODING"] = "utf-8"
        started = time.monotonic()
        try:
            completed = subprocess.run(
                [sys.executable, str(debug_path)],
                cwd=str(get_workspace()),
                env=env,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=30,
                check=False,
            )
            run_result = {
                "success": completed.returncode == 0,
                "returncode": completed.returncode,
                "timed_out": False,
                "elapsed_seconds": round(time.monotonic() - started, 3),
                "stdout": completed.stdout or "",
                "stderr": completed.stderr or "",
            }
        except subprocess.TimeoutExpired as exc:
            run_result = {
                "success": False,
                "returncode": None,
                "timed_out": True,
                "elapsed_seconds": round(time.monotonic() - started, 3),
                "stdout": str(exc.stdout or ""),
                "stderr": str(exc.stderr or ""),
            }
        debug_result: Dict[str, Any] = {}
        for line in reversed(str(run_result.get("stdout") or "").splitlines()):
            if not line.startswith("CODE_DEBUG_JSON="):
                continue
            try:
                parsed = json.loads(line[len("CODE_DEBUG_JSON="):])
                debug_result = parsed if isinstance(parsed, dict) else {}
            except Exception:
                debug_result = {}
            break
        static_result = debug_result.get("static") if isinstance(debug_result.get("static"), dict) else {}
        return _json({
            "success": bool(run_result.get("success") and static_result.get("syntax_ok")),
            "debug_file": created.get("debug_file"),
            "target_file": created.get("target_file"),
            "debug_result": debug_result,
            "debug_run": {
                "returncode": run_result.get("returncode"),
                "timed_out": run_result.get("timed_out", False),
                "elapsed_seconds": run_result.get("elapsed_seconds"),
                "stderr": run_result.get("stderr", ""),
            },
        })
    except Exception as exc:
        return _json({
            "success": False,
            "error": repr(exc),
        })


def _read_csv_info(path: Path, max_rows: int = 0) -> Dict[str, Any]:
    rows_preview = []
    row_count = 0
    fields: List[str] = []

    with path.open("r", encoding="utf-8-sig", newline="", errors="replace") as f:
        reader = csv.DictReader(f)
        fields = reader.fieldnames or []

        for row in reader:
            row_count += 1
            if max_rows and len(rows_preview) < max_rows:
                rows_preview.append(dict(row))

    return {
        "format": "csv",
        "rows": row_count,
        "fields": fields,
        "preview": rows_preview,
    }


def _read_json_info(path: Path, max_rows: int = 0) -> Dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8", errors="replace"))

    rows = 0
    fields: List[str] = []
    preview: Any = None

    if isinstance(data, list):
        rows = len(data)
        preview = data[:max_rows] if max_rows else None

        if data and isinstance(data[0], dict):
            fields = list(data[0].keys())

    elif isinstance(data, dict):
        list_key = None

        for key, value in data.items():
            if isinstance(value, list):
                list_key = key
                break

        if list_key:
            items = data[list_key]
            rows = len(items)
            preview = items[:max_rows] if max_rows else None

            if items and isinstance(items[0], dict):
                fields = list(items[0].keys())
        else:
            rows = 1
            fields = list(data.keys())
            preview = data if max_rows else None

    else:
        rows = 1
        preview = data if max_rows else None

    return {
        "format": "json",
        "rows": rows,
        "fields": fields,
        "preview": preview,
    }


def _read_xlsx_info(path: Path, max_rows: int = 0) -> Dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    values = ws.iter_rows(values_only=True)

    try:
        header = next(values)
    except StopIteration:
        return {
            "format": "xlsx",
            "rows": 0,
            "fields": [],
            "preview": [],
        }

    fields = [str(x) if x is not None else "" for x in header]

    row_count = 0
    preview = []

    for row in values:
        row_count += 1

        if max_rows and len(preview) < max_rows:
            item = {
                fields[i]: row[i] if i < len(row) else None
                for i in range(len(fields))
            }
            preview.append(item)

    return {
        "format": "xlsx",
        "rows": row_count,
        "fields": fields,
        "preview": preview,
    }


def _data_file_info(path: Path, max_rows: int = 0) -> Dict[str, Any]:
    suffix = path.suffix.lower()

    if suffix == ".csv":
        info = _read_csv_info(path, max_rows=max_rows)
    elif suffix == ".json":
        info = _read_json_info(path, max_rows=max_rows)
    elif suffix in {".xlsx", ".xlsm"}:
        info = _read_xlsx_info(path, max_rows=max_rows)
    else:
        raise ValueError("仅支持检查 .csv / .json / .xlsx / .xlsm 数据文件")

    base = _file_info(path)
    base.update(info)
    base["empty"] = info.get("rows", 0) == 0

    return base


def check_data_file(filename: str) -> str:
    """
    检查数据文件是否生成，并返回行数、字段名、文件大小等信息。
    支持 CSV、JSON、Excel。
    """
    try:
        path = _safe_path(filename)

        if not path.exists():
            return _json({
                "success": False,
                "exists": False,
                "error": "数据文件不存在",
                "filename": filename,
            })

        if not path.is_file():
            return _json({
                "success": False,
                "exists": False,
                "error": "目标不是文件",
                "filename": filename,
            })

        info = _data_file_info(path, max_rows=0)

        return _json({
            "success": True,
            "exists": True,
            "data_file": info,
        })

    except Exception as e:
        return _json({
            "success": False,
            "error": repr(e),
        })


def preview_data_file(filename: str, max_rows: int = 5) -> str:
    """
    预览数据文件前几行，并返回字段名和总行数。
    支持 CSV、JSON、Excel。
    """
    try:
        path = _safe_path(filename)

        if not path.exists():
            return _json({
                "success": False,
                "exists": False,
                "error": "数据文件不存在",
                "filename": filename,
            })

        max_rows = max(1, min(int(max_rows), 20))

        info = _data_file_info(path, max_rows=max_rows)

        return _json({
            "success": True,
            "exists": True,
            "data_file": info,
        })

    except Exception as e:
        return _json({
            "success": False,
            "error": repr(e),
        })
